from openpyxl import load_workbook 
from openpyxl import Workbook

#"과에서 데이터 가져오기"에서 가져온 데이터로 리스트 작성하기
wb=load_workbook("전체수강표.xlsx") 
ws=wb.active 

c_sub="전기" #선택한 학과
age=2 #선택한 학년
sub_data=[] #검색된 학과의 강좌들
one_sub=[] #한 강의제목에 해당하는 강좌들
tem_sub=None #일시적인 과목명(시간표 특징이 같은 과목명이 나열되어있으니)

#->선택한 학과이면서, 선택한 학년일 경우만 정보 추출
for row in ws.iter_rows(min_row=2): #min_row: 숫자에 해당하는 줄은 스킵해서 건너뜀
    if c_sub in str(row[13].value) and int(row[2].value)==age:

        #같은 교과목명끼리 묶어서 저장 [[교과1, (강의1), (강의2)...], [교과2, (강의3), (강의4)..], ..] 형식으로
        if str(row[5].value)==tem_sub:
            one_sub.append(tuple(row))
        else:
            sub_data.append(one_sub)
            one_sub=[]
            tem_sub=str(row[5].value)
            one_sub.append(str(row[5].value))
            one_sub.append(tuple(row))

sub_data.append(one_sub)
del sub_data[0]
print(sub_data)

#->같은 교과목에서 이제 시간을 구분해야해. 그중에 하나를 선택해야 하므로

"""
과목이 겹치면 안되잖아. 과목을 또 따로 분류시켜야 할듯한데. 아 학년이 있네. 
게다가 시간 구분은 어떻게 하지
"""



"""
보니깐 전체 수강학과는 무조건 포함이니깐 따로 만들어야 겠어. 근데 전체에서 중요한 것은 기교, 심교잖아. 
어떤 기교 심교 항목을 고를것인지 탭을 새로 만들어야 할 듯.
"""

"""
#기본 형식을 가져와서 작성하기
wb=load_workbook("timetable_basic.xlsx") 
ws=wb.active 

number=100 #나온 가지수를 시트를 표현해 늘리자. 

ws["A1"]="2021년 1학기 시간표 {}".format(number) #시간표 제목 

'''
과에서 데이터 가져오기 -> "전체데이터" 엑셀파일에서 수강학과 탭에서(N열) "전체"+"선택한 과"로 분리하기(따로 엑셀파일을 만들까 리스트로 만들까)
교수님/강의 선택하기 -> 선택한 강의만 따로 가져오기(무조건 추가해야 하는것), 리스트 변수로 튜플형식으로 넣어서 저장할까

'''

#저장
wb.save("test.xlsx")
wb.close()
"""